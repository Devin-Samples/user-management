"""Spreadsheet parser for users and orgs CSV/XLSX files.

Supports two formats:
1. **Legacy (action-based)**: CSV has an 'action' column (add/remove/update).
2. **Desired-state (sync)**: No 'action' column — CSV represents the desired state.

Auto-detection: if the header row contains an 'action' column the legacy parser
is used; otherwise the desired-state parser is used.

Also provides writer functions (:func:`write_orgs_csv`, :func:`write_users_csv`)
for emitting sync-format CSVs that can be round-tripped through the parsers.
"""
import csv
import os
import re
import warnings
from dataclasses import dataclass, field
from typing import IO

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------
class SpreadsheetValidationError(Exception):
    """Raised when spreadsheet data fails validation."""

    def __init__(self, row_num: int, field_name: str, message: str):
        self.row_num = row_num
        self.field_name = field_name
        super().__init__(f"Row {row_num}: {field_name} — {message}")


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------
@dataclass
class UserRow:
    email: str
    action: str  # "add" or "remove" (legacy); "sync" for desired-state
    enterprise_role: str = ""
    org_name: str = ""
    org_role: str = ""


@dataclass
class OrgRow:
    org_name: str
    action: str  # "add", "remove", "update" (legacy); "sync" for desired-state
    cycle_acu_limit: int | None = None
    session_acu_limit: int | None = None
    repos: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Validation helpers
# ---------------------------------------------------------------------------
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _validate_email(email: str, row_num: int) -> None:
    if not email or not email.strip():
        raise SpreadsheetValidationError(row_num, "email", "email is required")
    if not _EMAIL_RE.match(email.strip()):
        raise SpreadsheetValidationError(row_num, "email", f"invalid email format: {email!r}")


def _validate_user_action(action: str, row_num: int) -> None:
    if not action or not action.strip():
        raise SpreadsheetValidationError(row_num, "action", "action is required")
    if action.strip().lower() not in ("add", "remove"):
        raise SpreadsheetValidationError(
            row_num, "action", f"action must be 'add' or 'remove', got {action!r}"
        )


def _validate_org_action(action: str, row_num: int) -> None:
    if not action or not action.strip():
        raise SpreadsheetValidationError(row_num, "action", "action is required")
    if action.strip().lower() not in ("add", "remove", "update"):
        raise SpreadsheetValidationError(
            row_num, "action", f"action must be 'add', 'remove', or 'update', got {action!r}"
        )


def _parse_optional_int(value: str | int | None) -> int | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    return int(value)


def _parse_repos(value: str | int | None) -> list[str]:
    if not value or not str(value).strip():
        return []
    return [r.strip() for r in str(value).split(";") if r.strip()]


# ---------------------------------------------------------------------------
# Legacy (action-based) parsing
# ---------------------------------------------------------------------------
def _get_default_org_name() -> str:
    """Return the default org name from the ORG_NAME environment variable."""
    return os.environ.get("ORG_NAME", "")


def _parse_user_row(row: dict[str, str], row_num: int) -> UserRow:
    email = (row.get("email") or "").strip()
    action = (row.get("action") or "").strip().lower()
    enterprise_role = (row.get("enterprise_role") or "").strip()
    org_name = (row.get("org_name") or "").strip()
    org_role = (row.get("org_role") or "").strip()

    # Fall back to ORG_NAME env var if org_name column is empty
    if not org_name:
        org_name = _get_default_org_name()

    _validate_email(email, row_num)
    _validate_user_action(action, row_num)

    if org_name and not org_role:
        org_role = "org_member"
        warnings.warn(
            f"Row {row_num}: org_name set but org_role missing, defaulting to 'org_member'",
            stacklevel=2,
        )

    return UserRow(
        email=email,
        action=action,
        enterprise_role=enterprise_role,
        org_name=org_name,
        org_role=org_role,
    )


def parse_users_csv(file_obj: IO[str]) -> list[UserRow]:
    """Parse a users CSV from a file-like object (legacy action-based format)."""
    reader = csv.DictReader(file_obj)
    rows: list[UserRow] = []
    for i, row in enumerate(reader, start=2):  # row 1 is header
        rows.append(_parse_user_row(row, i))
    return rows


def _parse_org_row(row: dict[str, str | int | None], row_num: int) -> OrgRow:
    org_name = str(row.get("org_name") or "").strip()
    action = str(row.get("action") or "").strip().lower()

    if not org_name:
        org_name = _get_default_org_name()
    if not org_name:
        raise SpreadsheetValidationError(row_num, "org_name", "org_name is required")
    _validate_org_action(action, row_num)

    return OrgRow(
        org_name=org_name,
        action=action,
        cycle_acu_limit=_parse_optional_int(row.get("cycle_acu_limit")),
        session_acu_limit=_parse_optional_int(row.get("session_acu_limit")),
        repos=_parse_repos(row.get("repos")),
    )


def parse_orgs_csv(file_obj: IO[str]) -> list[OrgRow]:
    """Parse an orgs CSV from a file-like object (legacy action-based format)."""
    reader = csv.DictReader(file_obj)
    rows: list[OrgRow] = []
    for i, row in enumerate(reader, start=2):
        rows.append(_parse_org_row(row, i))
    return rows


# ---------------------------------------------------------------------------
# Desired-state (sync) parsing — no "action" column
# ---------------------------------------------------------------------------
def _parse_user_row_sync(row: dict[str, str], row_num: int) -> UserRow:
    """Parse a user row in desired-state format (no action column)."""
    email = (row.get("email") or "").strip()
    enterprise_role = (row.get("enterprise_role") or "").strip()
    org_name = (row.get("org_name") or "").strip()
    org_role = (row.get("org_role") or "").strip()

    _validate_email(email, row_num)

    if not enterprise_role:
        enterprise_role = "account_member"

    # Fall back to ORG_NAME env var if org_name column is empty
    if not org_name:
        org_name = _get_default_org_name()

    if org_name and not org_role:
        org_role = "org_member"
        warnings.warn(
            f"Row {row_num}: org_name set but org_role missing, defaulting to 'org_member'",
            stacklevel=2,
        )

    return UserRow(
        email=email,
        action="sync",
        enterprise_role=enterprise_role,
        org_name=org_name,
        org_role=org_role,
    )


def parse_users_csv_sync(file_obj: IO[str]) -> list[UserRow]:
    """Parse a desired-state users CSV (no action column)."""
    reader = csv.DictReader(file_obj)
    rows: list[UserRow] = []
    for i, row in enumerate(reader, start=2):
        rows.append(_parse_user_row_sync(row, i))
    return rows


def _parse_org_row_sync(row: dict[str, str | int | None], row_num: int) -> OrgRow:
    """Parse an org row in desired-state format (no action column)."""
    org_name = str(row.get("org_name") or "").strip()

    if not org_name:
        org_name = _get_default_org_name()
    if not org_name:
        raise SpreadsheetValidationError(row_num, "org_name", "org_name is required")

    return OrgRow(
        org_name=org_name,
        action="sync",
        cycle_acu_limit=_parse_optional_int(row.get("cycle_acu_limit")),
        session_acu_limit=_parse_optional_int(row.get("session_acu_limit")),
        repos=_parse_repos(row.get("repos")),
    )


def parse_orgs_csv_sync(file_obj: IO[str]) -> list[OrgRow]:
    """Parse a desired-state orgs CSV (no action column)."""
    reader = csv.DictReader(file_obj)
    rows: list[OrgRow] = []
    for i, row in enumerate(reader, start=2):
        rows.append(_parse_org_row_sync(row, i))
    return rows


# ---------------------------------------------------------------------------
# XLSX helpers
# ---------------------------------------------------------------------------
def _xlsx_to_dicts(path: str) -> list[dict[str, str | int | None]]:
    """Read an XLSX file and return a list of dicts (header → value)."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip().lower() if h else "" for h in next(rows_iter)]
    result: list[dict[str, str | int | None]] = []
    for row in rows_iter:
        d: dict[str, str | int | None] = {}
        for h, v in zip(headers, row):
            d[h] = str(v) if v is not None else ""
        result.append(d)
    wb.close()
    return result


# ---------------------------------------------------------------------------
# Auto-detecting file-level entry points
# ---------------------------------------------------------------------------
def _detect_has_action_column(path: str) -> bool:
    """Check whether a CSV/XLSX file has an 'action' column in its header."""
    if path.lower().endswith(".csv"):
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader, [])
            return "action" in [h.strip().lower() for h in header]
    elif path.lower().endswith(".xlsx"):
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return False
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        wb.close()
        return "action" in [str(h).strip().lower() for h in first_row if h]
    return False


def parse_users_file(path: str) -> list[UserRow]:
    """Parse a users file (CSV or XLSX). Auto-detects legacy vs desired-state format."""
    if path.lower().endswith(".csv"):
        has_action = _detect_has_action_column(path)
        with open(path, newline="", encoding="utf-8") as f:
            if has_action:
                return parse_users_csv(f)
            return parse_users_csv_sync(f)
    elif path.lower().endswith(".xlsx"):
        dicts = _xlsx_to_dicts(path)
        has_action = any("action" in d for d in dicts[:1]) if dicts else False
        # More reliable: check the keys
        if dicts and "action" in dicts[0]:
            rows: list[UserRow] = []
            for i, d in enumerate(dicts, start=2):
                str_d = {k: str(v) if v is not None else "" for k, v in d.items()}
                rows.append(_parse_user_row(str_d, i))
            return rows
        else:
            rows = []
            for i, d in enumerate(dicts, start=2):
                str_d = {k: str(v) if v is not None else "" for k, v in d.items()}
                rows.append(_parse_user_row_sync(str_d, i))
            return rows
    else:
        raise ValueError(f"Unsupported file extension: {path}. Use .csv or .xlsx")


def parse_orgs_file(path: str) -> list[OrgRow]:
    """Parse an orgs file (CSV or XLSX). Auto-detects legacy vs desired-state format."""
    if path.lower().endswith(".csv"):
        has_action = _detect_has_action_column(path)
        with open(path, newline="", encoding="utf-8") as f:
            if has_action:
                return parse_orgs_csv(f)
            return parse_orgs_csv_sync(f)
    elif path.lower().endswith(".xlsx"):
        dicts = _xlsx_to_dicts(path)
        if dicts and "action" in dicts[0]:
            rows: list[OrgRow] = []
            for i, d in enumerate(dicts, start=2):
                rows.append(_parse_org_row(d, i))
            return rows
        else:
            rows = []
            for i, d in enumerate(dicts, start=2):
                rows.append(_parse_org_row_sync(d, i))
            return rows
    else:
        raise ValueError(f"Unsupported file extension: {path}. Use .csv or .xlsx")


# ---------------------------------------------------------------------------
# Writers — emit sync-format CSVs suitable for round-tripping
# ---------------------------------------------------------------------------
ORGS_SYNC_HEADERS = ["org_name", "cycle_acu_limit", "session_acu_limit", "repos"]
USERS_SYNC_HEADERS = ["email", "enterprise_role", "org_name", "org_role"]


def write_orgs_csv(rows: list[OrgRow], file_obj: IO[str]) -> None:
    """Write a list of OrgRow objects as a sync-format CSV (no action column).

    The output is deterministic: rows are written in the order provided and
    empty ACU limit fields are written as empty strings. Repos are joined with
    semicolons to match the parser's expectations.
    """
    writer = csv.writer(file_obj)
    writer.writerow(ORGS_SYNC_HEADERS)
    for row in rows:
        writer.writerow([
            row.org_name,
            "" if row.cycle_acu_limit is None else str(row.cycle_acu_limit),
            "" if row.session_acu_limit is None else str(row.session_acu_limit),
            ";".join(row.repos),
        ])


def write_users_csv(rows: list[UserRow], file_obj: IO[str]) -> None:
    """Write a list of UserRow objects as a sync-format CSV (no action column).

    Each row represents either a standalone user (no org assignment) or a
    single (user, org) membership. Users with multiple org memberships are
    expected to appear on multiple rows.
    """
    writer = csv.writer(file_obj)
    writer.writerow(USERS_SYNC_HEADERS)
    for row in rows:
        writer.writerow([
            row.email,
            row.enterprise_role,
            row.org_name,
            row.org_role,
        ])
